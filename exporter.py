"""
exporter.py - 数据导出模块
支持 CSV / Excel（带格式）/ HTML 富文本报告
"""

import csv
import os
from datetime import datetime
from html import escape


COLUMNS = [
    ('title',        '标题'),
    ('platform',     '来源平台'),
    ('keywords',     '关键词'),
    ('publish_date', '发布日期'),
    ('crawl_time',   '采集时间'),
    ('snippet',      '摘要'),
    ('content',      '正文'),
    ('url',          '链接'),
]

COL_WIDTHS = {
    'title': 42, 'platform': 16, 'keywords': 22, 'publish_date': 14,
    'crawl_time': 20, 'snippet': 50, 'content': 80, 'url': 55,
}


# ──────────────────────────────────────────────
# CSV 导出
# ──────────────────────────────────────────────

def export_csv(results: list, filepath: str):
    """导出 CSV（UTF-8-BOM，直接用 Excel 打开不乱码）"""
    if not results:
        return
    fieldnames = [k for k, _ in COLUMNS]
    with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction='ignore')
        writer.writeheader()
        writer.writerows(results)


# ──────────────────────────────────────────────
# Excel 导出
# ──────────────────────────────────────────────

def export_excel(results: list, filepath: str):
    """导出带格式的 Excel 文件"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = '舆情数据'

    # ---- 表头样式 ----
    header_font = Font(bold=True, color='FFFFFF', size=11, name='微软雅黑')
    header_fill = PatternFill('solid', fgColor='2563EB')
    alt_fill    = PatternFill('solid', fgColor='EFF6FF')
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    top_wrap     = Alignment(vertical='top', wrap_text=True)
    top_nowrap   = Alignment(vertical='top')

    # 写表头
    for col_idx, (key, label) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font   = header_font
        cell.fill   = header_fill
        cell.alignment = center_align
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS.get(key, 20)

    ws.row_dimensions[1].height = 30

    # 写数据行
    for row_idx, item in enumerate(results, 2):
        for col_idx, (key, _) in enumerate(COLUMNS, 1):
            value = item.get(key, '') or ''
            # 超长正文截断（Excel 单格上限约 32767 字符）
            if key == 'content' and len(value) > 5000:
                value = value[:5000] + '\n…（内容已截断）'
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = (top_wrap if key in ('snippet', 'content') else top_nowrap)
            if row_idx % 2 == 0:
                cell.fill = alt_fill

        ws.row_dimensions[row_idx].height = 60

    # 冻结表头 + 自动筛选
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

    wb.save(filepath)


# ──────────────────────────────────────────────
# HTML 富文本报告导出
# ──────────────────────────────────────────────

def export_html(results: list, filepath: str):
    """导出可直接在浏览器查看的富文本 HTML 报告"""

    cards_html = ''
    for item in results:
        title    = escape(item.get('title', '无标题'))
        url      = escape(item.get('url', ''))
        platform = escape(item.get('platform', ''))
        keywords = escape(item.get('keywords', ''))
        pub_date = escape(item.get('publish_date', '') or item.get('crawl_time', ''))
        snippet  = escape(item.get('snippet', ''))
        content  = escape((item.get('content') or item.get('snippet') or '')[:800])
        content  = content.replace('\n', '<br>')

        cards_html += f'''
    <div class="card">
      <div class="card-title"><a href="{url}" target="_blank">{title}</a></div>
      <div class="card-meta">
        <span class="tag tag-platform">{platform}</span>
        <span class="tag tag-kw">{keywords}</span>
        <span class="date">{pub_date}</span>
      </div>
      <div class="card-content">{content}</div>
      <div class="card-url"><a href="{url}" target="_blank">{url}</a></div>
    </div>'''

    now_str = datetime.now().strftime('%Y年%m月%d日 %H:%M')
    total   = len(results)

    html = f'''<!DOCTYPE html>
<html lang="zh-CN">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>舆情监控报告 · {now_str}</title>
  <style>
    *, *::before, *::after {{ box-sizing: border-box; margin: 0; padding: 0; }}
    body {{
      font-family: -apple-system, BlinkMacSystemFont, "PingFang SC", "Microsoft YaHei",
                   "Segoe UI", sans-serif;
      background: #f0f4f8; color: #1e293b; padding: 32px 16px;
    }}
    .page-wrapper {{ max-width: 960px; margin: 0 auto; }}
    header {{ margin-bottom: 28px; }}
    header h1 {{ font-size: 24px; color: #1d4ed8; font-weight: 700; }}
    header p  {{ color: #64748b; margin-top: 4px; font-size: 14px; }}
    .summary {{
      background: #fff; border-radius: 10px; padding: 16px 20px;
      margin-bottom: 24px; display: flex; gap: 24px; flex-wrap: wrap;
      box-shadow: 0 1px 3px rgba(0,0,0,.08);
    }}
    .summary-item {{ display: flex; flex-direction: column; }}
    .summary-item .label {{ font-size: 12px; color: #94a3b8; }}
    .summary-item .value {{ font-size: 20px; font-weight: 700; color: #1d4ed8; }}
    .card {{
      background: #fff; border-radius: 12px; padding: 20px 24px;
      margin-bottom: 16px; box-shadow: 0 1px 4px rgba(0,0,0,.08);
      transition: box-shadow .2s;
    }}
    .card:hover {{ box-shadow: 0 4px 12px rgba(0,0,0,.12); }}
    .card-title {{ font-size: 16px; font-weight: 600; margin-bottom: 8px; }}
    .card-title a {{ color: #1d4ed8; text-decoration: none; }}
    .card-title a:hover {{ text-decoration: underline; }}
    .card-meta {{ display: flex; gap: 8px; align-items: center; flex-wrap: wrap; margin-bottom: 12px; }}
    .tag {{
      padding: 2px 10px; border-radius: 20px; font-size: 12px; font-weight: 500;
    }}
    .tag-platform {{ background: #dbeafe; color: #1d4ed8; }}
    .tag-kw       {{ background: #fce7f3; color: #be185d; }}
    .date {{ color: #94a3b8; font-size: 12px; }}
    .card-content {{
      color: #475569; font-size: 14px; line-height: 1.75;
      border-left: 3px solid #bfdbfe; padding-left: 14px;
      margin-bottom: 12px;
    }}
    .card-url {{ font-size: 12px; color: #94a3b8; word-break: break-all; }}
    .card-url a {{ color: #94a3b8; }}
    footer {{ text-align: center; color: #94a3b8; font-size: 12px; margin-top: 32px; }}
  </style>
</head>
<body>
  <div class="page-wrapper">
    <header>
      <h1>舆情监控报告</h1>
      <p>生成时间：{now_str}</p>
    </header>
    <div class="summary">
      <div class="summary-item">
        <span class="label">数据总量</span>
        <span class="value">{total}</span>
      </div>
    </div>
    {cards_html}
    <footer>由舆情监控工具自动生成 · {now_str}</footer>
  </div>
</body>
</html>'''

    with open(filepath, 'w', encoding='utf-8') as f:
        f.write(html)
